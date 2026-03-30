"""
config.py — Static configuration for the Y-Strainer Pressure Drop Calculator.

All physical constants and lookup tables are stored here.
Tabular data is loaded from 'Pr drop cal data for Y strainer.xlsx' at import time.

Sources:
  - Perry's Chemical Engineers' Handbook, pg 5-40  -> C_VALUE_TABLE
  - ASTM E2016-15                                  -> MESH_DATA
  - ASTM E674-12                                   -> PERF_SHEET_DATA
  - Manufacturer strainer catalogue                -> STRAINER_DATA
"""

from pathlib import Path
import openpyxl

# ── Physical constant ──────────────────────────────────────────────────────────
GRAVITY_CM_S2: float = 981.0          # gravitational acceleration  (cm/s²)

# Unit conversion factors
CM3_PER_M3: float = 1_000_000.0       # 1 m³ = 1,000,000 cm³
SECONDS_PER_HOUR: float = 3600.0

# Pressure unit conversion constants
CM_WC_TO_IN_WC: float = 1 / 2.54     # 1 cm WC = 1/2.54 in WC
CM_WC_TO_KG_CM2: float = 1 / 1000.0  # 1 cm WC = 0.001 kg/cm²

# ── C value table — Perry's Chemical Engineers' Handbook, pg 5-40 ─────────────
# Format: (Re_lower, Re_upper_or_None, C)
# For Re < 21: formula  C = sqrt(Re) / 10  is used (derived from software behaviour)
C_VALUE_TABLE: list = [
    (5500, None, 1.50), (5001, 5499, 1.49), (4990, 5000, 1.48), (4500, 4989, 1.47),
    (4000, 4499, 1.46), (3500, 3999, 1.45), (3300, 3499, 1.44), (2907, 3299, 1.43),
    (2666, 2906, 1.42), (2659, 2665, 1.42), (2651, 2658, 1.42), (2504, 2650, 1.41),
    (2200, 2503, 1.40), (2093, 2199, 1.39), (1914, 2092, 1.38), (1841, 1913, 1.37),
    (1800, 1840, 1.36), (1700, 1799, 1.35), (1600, 1699, 1.34), (1500, 1599, 1.33),
    (1400, 1499, 1.32), (1250, 1399, 1.31), (1150, 1249, 1.30), (950,  1149, 1.29),
    (860,   949, 1.28), (830,   859, 1.27), (810,   829, 1.26), (716,   809, 1.25),
    (700,   715, 1.24), (675,   699, 1.23), (650,   674, 1.22), (625,   649, 1.21),
    (600,   624, 1.20), (575,   599, 1.19), (550,   574, 1.18), (525,   549, 1.17),
    (500,   524, 1.16), (490,   499, 1.15), (480,   489, 1.14), (470,   479, 1.13),
    (460,   469, 1.12), (450,   459, 1.11), (440,   449, 1.10), (430,   439, 1.09),
    (410,   429, 1.08), (400,   409, 1.07), (390,   399, 1.06), (376,   389, 1.05),
    (350,   375, 1.04), (340,   349, 1.03), (320,   339, 1.02), (301,   319, 1.01),
    (201,   300, 1.00), (175,   200, 0.95), (150,   174, 0.90), (125,   149, 0.85),
    (100,   124, 0.80), (90,     99, 0.77), (80,     89, 0.75), (70,     79, 0.73),
    (60,     69, 0.70), (50,     59, 0.65), (40,     49, 0.60), (39,     40, 0.59),
    (38,     39, 0.58), (37,     38, 0.57), (36,     37, 0.56), (35,     36, 0.55),
    (34,     35, 0.54), (33,     34, 0.53), (32,     33, 0.52), (31,     32, 0.51),
    (30,     31, 0.50), (29,     30, 0.49), (28,     29, 0.48), (27,     28, 0.47),
    (26,     27, 0.46), (21,     25, 0.45),
]

# ── Excel-backed lookup tables ─────────────────────────────────────────────────
_EXCEL_PATH = Path(__file__).parent / "Pr drop cal data for Y strainer.xlsx"


def _load_tables() -> tuple:
    """Load all lookup tables from the Excel reference file.

    Returns:
        mesh_data       : {(mesh_int, swg_int): (opening_mm, open_area_pct)}
        perf_sheet_data : {description_upper_stripped: open_area_pct}
        strainer_data   : {(model_upper, nps_str): (pipe_OD_mm, screen_D_mm, screen_L_mm)}
        pipe_data       : {(nps_str, schedule_str): id_mm}
    """
    wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True, read_only=True)

    # Mesh Data — ASTM E2016-15
    # Columns: mesh, swg, openSize(mm), OpenArea(%)
    mesh_data: dict = {}
    for row in wb["Mesh Data"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            mesh_data[(int(row[0]), int(row[1]))] = (float(row[2]), float(row[3]))

    # Perf Sheet — ASTM E674-12
    # Columns: SrNo, Perforation(description), PercentageOpenP(%)
    perf_sheet_data: dict = {}
    for row in wb["Perf Sheet"].iter_rows(min_row=2, values_only=True):
        if row[1] is not None:
            perf_sheet_data[str(row[1]).strip().upper()] = float(row[2])

    # Strainer Data — manufacturer dimensional catalogue
    # Columns: SType, SrNo, NPS, OD(nominal pipe mm), D(screen OD mm), L(screen length mm)
    strainer_data: dict = {}
    for row in wb["Strainer data"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[2] is not None:
            key = (str(row[0]).strip().upper(), str(row[2]).strip())
            strainer_data[key] = (float(row[3]), float(row[4]), float(row[5]))

    # Pipe Sizes — standard pipe schedule dimensions
    # Columns: sl_no, NPS_inch, Nominal_bore, OD, SchNoCS1, SchNoCS2, ID, DN, ...
    pipe_data: dict = {}
    pipe_nps_data: dict = {}   # {nps_inch_str: dn_mm} — first unique occurrence wins
    pipe_schedule_data: dict = {}  # {nps_inch_str: [(schedule_str, id_mm), ...]}
    _SCH_SKIP = ("-", "\u2013", "\u2014")  # hyphens / dashes that mean "no named schedule"
    for row in wb["Pipe sizes"].iter_rows(min_row=2, values_only=True):
        nps  = str(row[1]).strip() if row[1] else None
        nb   = str(row[2]).strip() if row[2] else None
        sch1 = str(row[4]).strip() if row[4] else ""   # SchNoCS1 — named (STD, XS, XXS)
        sch  = str(row[5]).strip() if row[5] else ""   # SchNoCS2 — numeric (5, 10, 40 …)
        id_val = row[6]
        dn_val = row[7]
        if id_val is not None:
            id_mm = float(id_val)
            if nps:
                pipe_data[(nps, sch)] = id_mm
                if sch1 and sch1 not in _SCH_SKIP:
                    pipe_data[(nps, sch1)] = id_mm
            if nb:
                pipe_data[(nb, sch)] = id_mm
                if sch1 and sch1 not in _SCH_SKIP:
                    pipe_data[(nb, sch1)] = id_mm
            # Build per-NPS schedule list for Basket strainer Schedule dropdown
            if nps:
                if nps not in pipe_schedule_data:
                    pipe_schedule_data[nps] = []
                if sch and sch not in _SCH_SKIP:
                    pipe_schedule_data[nps].append((sch, id_mm))
                if sch1 and sch1 not in _SCH_SKIP:
                    pipe_schedule_data[nps].append((sch1, id_mm))
        if nps and dn_val is not None and nps not in pipe_nps_data:
            try:
                pipe_nps_data[nps] = float(dn_val)
            except (ValueError, TypeError):
                pass  # skip rows where DN is not a plain number

    # Pr Class — pressure rating classes
    # Columns: Rating, RatingNo
    pr_class_data: list = []
    for row in wb["Pr Class"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try:
                pr_class_data.append(int(row[0]))
            except (ValueError, TypeError):
                pass
    pr_class_data.sort()

    wb.close()
    return mesh_data, perf_sheet_data, strainer_data, pipe_data, pipe_nps_data, pr_class_data, pipe_schedule_data


try:
    MESH_DATA, PERF_SHEET_DATA, STRAINER_DATA, PIPE_DATA, PIPE_NPS_DATA, PR_CLASS_DATA, PIPE_SCHEDULE_DATA = _load_tables()
except Exception as _exc:
    raise RuntimeError(
        f"Failed to load reference tables from '{_EXCEL_PATH}': {_exc}"
    ) from _exc
